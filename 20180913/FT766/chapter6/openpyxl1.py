# 從openpyxl裡匯入必要的函式（load_workbook）
from openpyxl import load_workbook
import numpy as np  # 也匯入NumPy

# 開啟WorkBook
wb = load_workbook(filename='Sample1.xlsx', read_only=True,
                   data_only=True, use_iterators=True)
# 以名稱指定WorkSheet
ws = wb['溫度變化']

# 預先生成用來儲存資料的NumPy的ndarray
Nrow = 11
time_vec = np.zeros(Nrow)
temp_vec = np.zeros(Nrow)

# 資料的讀取
for i, row in enumerate(ws.iter_rows(row_offset=1)):
    time_vec[i] = row[0].value
    temp_vec[i] = row[1].value
